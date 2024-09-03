import shutil
import pandas as pd
import openpyxl
from datetime import date
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Define custom fill colors
fill_colors = {
    'green': PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid'),
    'red': PatternFill(start_color='ff9c9c', end_color='ff9c9c', fill_type='solid'),
    'orange': PatternFill(start_color='fcba8d', end_color='fcba8d', fill_type='solid'),
    'yellow': PatternFill(start_color='f7e099', end_color='f7e099', fill_type='solid'),
    'blue': PatternFill(start_color='3A5FCD', end_color='3A5FCD', fill_type='solid')
}


# Fill excel-file based on color code
def fill_excels(methods, path_excel_filled, sheet_name):
    # Extract individual methods from the provided list
    method1 = methods[0]
    method2 = methods[1]

    # Load the filled excel workbook using openpyxl
    file_excel_filled = openpyxl.load_workbook(path_excel_filled)
    cur_wb = file_excel_filled[sheet_name]

    # Get the maximum number of rows and columns for iteration purposes
    maxlength_row = cur_wb.max_row
    maxlength_column = cur_wb.max_column
    minlength_column = cur_wb.min_column

    # Iterate over the columns starting from the second column
    for col in cur_wb.iter_cols(min_col=2, max_col=maxlength_column-6):
        # Initialize counters for each color
        cur_counter_orange, cur_counter_yellow, cur_counter_green, cur_counter_red = 0, 0, 0, 0

        # For each cell in the column
        for cell in col:
            # Extract method names from adjacent cells
            cur_cell_method1 = cur_wb.cell(row=cell.row, column=1).value
            cur_cell_method2 = cur_wb.cell(row=cell.row + 3, column=1).value

            # Conditions to check if methods need replacements and fill corresponding rows with blue
            if len(str(cur_cell_method1)) > 4 and len(str(cur_cell_method2)) > 4:
                cell_method1_method2 = cur_cell_method2.replace(method2, method1)
                cell_method1_method2_2 = cur_cell_method2.replace(method1, method2)

                if cur_cell_method1[-8:] == method1 and cur_cell_method1 != cell_method1_method2:
                    cur_wb = fill_row_colour(cur_wb, cell.row, maxlength_column, fill_colors['blue'])

                elif cur_cell_method1[-8:] == method2 and cur_cell_method1 != cell_method1_method2_2:
                    cur_wb = fill_row_colour(cur_wb, cell.row, maxlength_column, fill_colors['blue'])

                else:
                    if cur_wb.cell(row=cell.row, column=1).value[-11:] == 'DP_' + method1:
                        if cur_wb.cell(row=cell.row, column=cell.column).value is None and not cur_wb.cell(
                                row=cell.row + 3, column=cell.column).value is None:
                            cur_counter_orange = cur_counter_orange + 1
                            for num in range(0, 6):
                                cur_wb.cell(row=cell.row + num, column=cell.column).fill = fill_colors['orange']

                        elif cur_wb.cell(row=cell.row, column=cell.column).value is not None and cur_wb.cell(
                                row=cell.row + 3, column=cell.column).value is None:
                            cur_counter_yellow = cur_counter_yellow + 1
                            for num in range(0, 6):
                                cur_wb.cell(row=cell.row + num, column=cell.column).fill = fill_colors['yellow']

                        elif cur_wb.cell(row=cell.row, column=cell.column).value is not None and cur_wb.cell(
                                row=cell.row + 3, column=cell.column).value is not None:
                            try:
                                curEntry_method2 = ''.join(sorted(cur_wb.cell(row=cell.row + 4, column=cell.column).value)
                                                           + sorted(cur_wb.cell(row=cell.row + 5, column=cell.column).value))
                            except: pass

                            if len(''.join(sorted(cur_wb.cell(row=cell.row + 1, column=cell.column).value))) == 3:
                                curEntry_method1 = ''.join(
                                    sorted(cur_wb.cell(row=cell.row + 1, column=cell.column).value) + sorted(
                                        cur_wb.cell(row=cell.row + 2, column=cell.column).value))
                                if sorted(curEntry_method1) == sorted(curEntry_method2):
                                    x = 0
                                else:
                                    curEntry_method1 = sorted(cur_wb.cell(row=cell.row + 1, column=cell.column).value)[
                                                            1] + sorted(
                                        cur_wb.cell(row=cell.row + 2, column=cell.column).value)[2]
                            else:
                                curEntry_method1 = ''.join(
                                    sorted(cur_wb.cell(row=cell.row + 1, column=cell.column).value) + sorted(
                                        cur_wb.cell(row=cell.row + 2, column=cell.column).value))

                            curEntry_method1 = sorted(curEntry_method1)
                            curEntry_method2 = sorted(curEntry_method2)

                            if curEntry_method1 == curEntry_method2:
                                cur_counter_green = cur_counter_green + 1
                                for num in range(0, 6):
                                    cur_wb.cell(row=cell.row + num, column=cell.column).fill = fill_colors['green']
                            elif curEntry_method1 != curEntry_method2:
                                cur_counter_red = cur_counter_red + 1
                                for num in range(0, 6):
                                    cur_wb.cell(row=cell.row + num, column=cell.column).fill = fill_colors['red']

            # Creating a bold font style
            bold_font = Font(bold=True)
            # Update the counters in the last 4 rows
            # fill countered green
            cur_wb.cell(row=maxlength_row+1, column=minlength_column).value = 'Number of matching results per variant position (green)'
            cur_wb.cell(row=maxlength_row + 1, column=minlength_column).font = bold_font
            cur_wb.cell(row=maxlength_row + 1, column=cell.column).value = cur_counter_green
            cur_wb.cell(row=maxlength_row + 1, column=cell.column).fill = fill_colors['green']

            # fill countered red
            cur_wb.cell(row=maxlength_row + 2, column=minlength_column).value = 'Number of mismatching results per variant position (red)'
            cur_wb.cell(row=maxlength_row + 2, column=minlength_column).font = bold_font
            cur_wb.cell(row=maxlength_row + 2, column=cell.column).value = cur_counter_red
            cur_wb.cell(row=maxlength_row + 2, column=cell.column).fill = fill_colors['red']

            # fill countered orange
            cur_wb.cell(row=maxlength_row + 3, column=minlength_column).value = 'Number of variants only detected by the first method (orange)'
            cur_wb.cell(row=maxlength_row + 3, column=minlength_column).font = bold_font
            cur_wb.cell(row=maxlength_row + 3, column=cell.column).value = cur_counter_orange
            cur_wb.cell(row=maxlength_row + 3, column=cell.column).fill = fill_colors['orange']

            # fill countered yellow
            cur_wb.cell(row=maxlength_row + 4, column=minlength_column).value = 'Number of variants only detected by the second method (yellow)'
            cur_wb.cell(row=maxlength_row + 4, column=minlength_column).font = bold_font
            cur_wb.cell(row=maxlength_row + 4, column=cell.column).value = cur_counter_yellow
            cur_wb.cell(row=maxlength_row + 4, column=cell.column).fill = fill_colors['yellow']

    # Save the modified workbook and close it
    file_excel_filled.save(path_excel_filled)
    file_excel_filled.close()
    return

# Fill rows with blue color, if only one of them is present, e.g. method2-Data for one sample is missing
def fill_row_colour(cur_wb, cur_row, max_length_column, blue_fill):
    for curCol in range(1, max_length_column):
        cur_wb.cell(row=cur_row, column=curCol).fill = blue_fill
    return cur_wb


def color_count(methods, sheet_name, excel_path):
    method1 = methods[0]
    method2 = methods[1]

    # Load the workbook and get the specified sheet
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    # Get the maximum column and row numbers for the sheet
    max_column = ws.max_column
    max_row = ws.max_row

    # Define a dictionary to map color codes to their corresponding names
    color_codes = {
        '00e2efda': 'green',
        '00ff9c9c': 'red',
        '00fcba8d': 'orange',
        '00f7e099': 'yellow'
    }

    # Creating a bold font style
    bold_font = Font(bold=True)

    # Setting values and making them bold
    ws.cell(row=1, column=max_column + 1).value = 'number of matching results per sample (green)'
    ws.cell(row=1, column=max_column + 1).font = bold_font

    ws.cell(row=1, column=max_column + 2).value = 'number of mismatching results per sample (red)'
    ws.cell(row=1, column=max_column + 2).font = bold_font

    ws.cell(row=1, column=max_column + 3).value = 'only the first method detected a haplotype per sample (orange)'
    ws.cell(row=1, column=max_column + 3).font = bold_font

    ws.cell(row=1, column=max_column + 4).value = 'only the second method detected a haplotype per sample (yellow)'
    ws.cell(row=1, column=max_column + 4).font = bold_font

    # Iterate through the rows in the sheet, starting from the third column and excluding the last 10 rows
    for row in ws.iter_rows(min_col=2, max_row=max_row -10):
        # Initialize a dictionary to store the count of each color in the row
        color_counts = {'green': 0, 'red': 0, 'orange': 0, 'yellow': 0}

        # Update color_counts dictionary based on bgColor or fgColor
        for cell in row:
            bg_color = cell.fill.bgColor.index
            fg_color = cell.fill.fgColor.index
            color = color_codes.get(bg_color) or color_codes.get(fg_color)
            if color:
                color_counts[color] += 1

        # Update the counts in the worksheet if the cell value ends with 'DP_' method1
        try:
            if ws.cell(row=row[0].row, column=1).value[-11:] == 'DP_' + method1:
                # Iterate through the fill_colors dictionary and update the worksheet with the color counts
                for i, (color, fill) in enumerate(fill_colors.items()):
                    if color != 'blue':
                        ws.cell(row=row[0].row, column=max_column + i + 1).value = color_counts[color]
                        ws.cell(row=row[0].row, column=max_column + i + 1).fill = fill
        except:
            pass

    # Save the modified workbook and close it
    wb.save(filename=excel_path)
    wb.close()


def format_excel_sheet(file_name_colored, sheet_name, cur_method):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(file_name_colored)
    sheet = workbook[sheet_name]

    # Define styles for lines
    thin_side = Side(style='thin')
    thin_border = Border(bottom=thin_side)
    double_border = Border(bottom=Side(style='double'))

    # Define alignment and font styles
    alignment = Alignment(horizontal='left')
    font = Font(name='Calibri', size=11)

    # Define style for the first 4 rows
    bold_font = Font(bold=True, name='Calibri')
    fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    #  Format the first 4 rows
    for row in sheet.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.fill = fill
            cell.alignment = alignment
            cell.font = font

    # Fixate the first 4 rows
    sheet.freeze_panes = 'A5'

    # Set the width of the first column
    sheet.column_dimensions['A'].width = 45

    # Iterate through rows to find barcode sections and draw horizontal lines
    last_row = sheet.max_row
    for row_idx, row in enumerate(sheet.iter_rows(), start=5):
        for cell in row:
            # Apply alignment and font styles to each cell
            cell.alignment = alignment
            cell.font = font

        first_cell_value = str(row[0].value or '')
        if first_cell_value.startswith('bc') and first_cell_value.endswith('_H2_' + cur_method):
            border_style = thin_border if row_idx != last_row else double_border
            for cell in row:
                current_border = cell.border
                new_border = Border(left=current_border.left,
                                    right=current_border.right,
                                    top=current_border.top,
                                    bottom=border_style.bottom)
                cell.border = new_border

        #  Format the first 4 rows
    for row in sheet.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.font = bold_font


    # Save the modified workbook
    workbook.save(file_name_colored)


def format_excel_sheet_haplotypes(file_name_colored, sheet_name):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(file_name_colored)
    sheet = workbook[sheet_name]


    # Define alignment and font styles
    alignment = Alignment(horizontal='left')
    font = Font(name='Calibri', size=11)

    # Style for the first 4 rows
    bold_font = Font(bold=True, name='Calibri')
    fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    # Format first row
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = fill
            cell.alignment = alignment
            cell.font = font

    # Fixate the first row
    sheet.freeze_panes = 'A2'
    # Set the width of the first three columns
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    for row_idx, row in enumerate(sheet.iter_rows(), start=2):
        for cell in row:
            # Apply alignment and font styles to each cell
            cell.alignment = alignment
            cell.font = font

    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = bold_font

    # Save the modified workbook
    workbook.save(file_name_colored)




# Define the settings to fill the rows and colors
def fill_row_and_color_count_setting(amplicon_data, methods, file_excel_analyzed, path_output):

    print('## Color Count Setting ##')

    # Get the current date
    today = date.today()

    file_name_colored = path_output + '/' + str(today) + '_step_4_dataset_color_coded.xlsx'

    # Copy the source file to the destination file
    shutil.copy(file_excel_analyzed, file_name_colored)

    # Loop through the sheet names and perform fill_excels and color_count functions
    for sheet_name in amplicon_data.keys():
        print('Coloring: ' + str(sheet_name))
        fill_excels(methods, file_name_colored, sheet_name)
        color_count(methods, sheet_name, file_name_colored)
        format_excel_sheet(file_name_colored, sheet_name, methods[1])
        format_excel_sheet_haplotypes(file_name_colored, sheet_name + '_haplotypes_by_methods')
        format_excel_sheet_haplotypes(file_name_colored, sheet_name + '_alleles_by_methods')
